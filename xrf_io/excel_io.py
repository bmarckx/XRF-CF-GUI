"""
Excel read/write for XRF Correction Factor projects.

Per-sheet layout (visible rows on each calibration/analysis sheet):
  Row 1  : title
  Row 2  : "Diameter (cm):"  | <value>
  Row 3  : "Area (cm²):"      | <formula = PI()*(B2/2)^2>
  Row 4  : "Input mode:"      | "mass" or "loading"
  Row 5  : (calibration) "Correction Factor (avg):" | <formula>
           (analysis) per-element CF labels and reference formulas
  Row 6  : (analysis only) CF values per element column
  Row 7  : blank
  Row 8  : header row
  Row 9+ : sample data

A hidden "_xrf_meta" sheet stores project-level metadata.
"""

import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.project import (
    Project, CalibrationSheet, CalibrationSample, AnalysisSheet, AnalysisSample,
    INPUT_MASS, INPUT_LOADING,
)


_HDR_FILL   = PatternFill("solid", fgColor="4472C4")
_HDR_FONT   = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=12)
_CFG_FILL   = PatternFill("solid", fgColor="FFF2CC")
_CF_FILL    = PatternFill("solid", fgColor="D9E1F2")
_THIN       = Side(border_style="thin", color="000000")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ═══════════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════════

def save_project(project: Project, path: str) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    meta = wb.create_sheet("_xrf_meta")
    meta["A1"] = "project_name"; meta["B1"] = project.name
    meta["A2"] = "diameter_cm";  meta["B2"] = project.diameter_cm
    meta.sheet_state = "hidden"

    for cs in project.calibration_sheets:
        _write_calibration_sheet(wb, cs, project)

    for ans in project.analysis_sheets:
        _write_analysis_sheet(wb, ans, project)

    wb.save(path)


def _write_config_rows(ws, sheet, project):
    """Rows 2–4: diameter, area, input mode."""
    ws["A2"] = "Diameter (cm):"
    ws["B2"] = sheet.effective_diameter(project)
    ws["A2"].fill = _CFG_FILL; ws["B2"].fill = _CFG_FILL

    ws["A3"] = "Area (cm²):"
    ws["B3"] = "=PI()*(B2/2)^2"
    ws["B3"].number_format = "0.0000"
    ws["A3"].fill = _CFG_FILL; ws["B3"].fill = _CFG_FILL

    ws["A4"] = "Input mode:"
    ws["B4"] = sheet.input_mode  # "mass" or "loading"
    ws["A4"].fill = _CFG_FILL; ws["B4"].fill = _CFG_FILL


def _write_calibration_sheet(wb, cs: CalibrationSheet, project: Project) -> None:
    title = f"{cs.element} on {cs.substrate}" if cs.substrate else cs.element
    ws = wb.create_sheet(title[:31])

    ws["A1"] = f"{cs.element} on {cs.substrate} — XRF correction factor"
    ws["A1"].font = _TITLE_FONT

    _write_config_rows(ws, cs, project)

    ws["A5"] = "Correction Factor (avg):"
    n = max(len(cs.samples), 1)
    ws["B5"] = f"=AVERAGE(E9:E{8 + n})"
    ws["B5"].number_format = "0.0000"
    ws["A5"].fill = _CF_FILL; ws["B5"].fill = _CF_FILL

    # row 7 blank, row 8 headers
    mass_col_lbl = "Mass Loading (mg/cm²)" if cs.input_mode == INPUT_LOADING else "Mass (mg)"
    unc_col_lbl  = "σ_loading (mg/cm²)"     if cs.input_mode == INPUT_LOADING else "σ_mass (mg)"
    headers = ["Sample ID", mass_col_lbl, "Mass Loading (mg/cm²)",
               "XRF Loading (mg/cm²)", "Corr. Factor",
               "Corr. Loading (mg/cm²)", "% Error", unc_col_lbl]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=8, column=col, value=h)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _BORDER

    # Mass loading formula references B3 (area) for mass mode, or just B for loading mode
    for row_i, s in enumerate(cs.samples, 9):
        if cs.input_mode == INPUT_LOADING:
            mass_loading_formula = f"=B{row_i}"
        else:
            mass_loading_formula = f"=B{row_i}/$B$3"
        col_vals = [
            s.sample_id,
            s.mass_mg,
            mass_loading_formula,
            s.xrf_loading,
            f"=C{row_i}/D{row_i}",
            f"=D{row_i}*$B$5",
            f"=ABS(F{row_i}-C{row_i})/C{row_i}*100",
            s.mass_uncertainty,
        ]
        for col_i, v in enumerate(col_vals, 1):
            c = ws.cell(row=row_i, column=col_i, value=v)
            c.border = _BORDER
            if col_i in (3, 4, 5, 6): c.number_format = "0.0000"
            if col_i == 7: c.number_format = "0.00"

    _autofit(ws)


def _write_analysis_sheet(wb, ans: AnalysisSheet, project: Project) -> None:
    ws = wb.create_sheet(ans.name[:31])
    ws["A1"] = f"{ans.name} — XRF correction validation"
    ws["A1"].font = _TITLE_FONT

    _write_config_rows(ws, ans, project)

    # Row 5: CF labels, row 6: CF reference formulas to calibration sheets
    ws["A5"] = "Correction Factors:"; ws["A5"].fill = _CF_FILL
    for col_i, el in enumerate(ans.elements, 2):
        cs = project.calibration_by_element(el)
        ws.cell(row=5, column=col_i, value=f"{el} CF:").fill = _CF_FILL
        ref = f"='{cs.element + (' on ' + cs.substrate if cs.substrate else '')}'!B5" if cs else 1.0
        c = ws.cell(row=6, column=col_i, value=ref)
        c.fill = _CF_FILL; c.number_format = "0.0000"

    # Row 8 headers
    mass_col_lbl = "Mass Loading (mg/cm²)" if ans.input_mode == INPUT_LOADING else "Mass (mg)"
    fixed_pre  = ["Sample ID", mass_col_lbl, "Mass Loading (mg/cm²)"]
    xrf_hdrs   = [f"XRF {el} (mg/cm²)" for el in ans.elements]
    fixed_post = ["XRF Total (mg/cm²)", "Corr. Total (mg/cm²)", "% Error", "σ_corr (mg/cm²)", "Notes"]
    headers    = fixed_pre + xrf_hdrs + fixed_post
    hdr_row    = 8

    for col_i, h in enumerate(headers, 1):
        c = ws.cell(row=hdr_row, column=col_i, value=h)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _BORDER

    n_pre = len(fixed_pre)
    n_xrf = len(ans.elements)
    xrf_start_col = n_pre + 1
    xrf_end_col   = n_pre + n_xrf
    total_col     = xrf_end_col + 1
    corr_col      = total_col + 1
    err_col       = corr_col + 1
    sig_col       = err_col + 1
    notes_col     = sig_col + 1

    for row_i, s in enumerate(ans.samples, hdr_row + 1):
        ws.cell(row=row_i, column=1, value=s.sample_id).border = _BORDER
        ws.cell(row=row_i, column=2, value=s.mass_mg).border = _BORDER

        if ans.input_mode == INPUT_LOADING:
            ml_formula = f"=B{row_i}"
        else:
            ml_formula = f"=B{row_i}/$B$3"
        c = ws.cell(row=row_i, column=3, value=ml_formula); c.border = _BORDER
        c.number_format = "0.0000"

        for el_i, el in enumerate(ans.elements):
            col_i = xrf_start_col + el_i
            c = ws.cell(row=row_i, column=col_i, value=s.xrf_loadings.get(el, ""))
            c.border = _BORDER; c.number_format = "0.0000"

        xrf_range = f"{get_column_letter(xrf_start_col)}{row_i}:{get_column_letter(xrf_end_col)}{row_i}"
        c = ws.cell(row=row_i, column=total_col, value=f"=SUM({xrf_range})")
        c.border = _BORDER; c.number_format = "0.0000"

        terms = []
        for el_i in range(len(ans.elements)):
            xrf_letter = get_column_letter(xrf_start_col + el_i)
            cf_letter  = get_column_letter(2 + el_i)   # row 6 holds CF values
            terms.append(f"{xrf_letter}{row_i}*{cf_letter}6")
        c = ws.cell(row=row_i, column=corr_col, value="=" + "+".join(terms))
        c.border = _BORDER; c.number_format = "0.0000"

        corr_letter = get_column_letter(corr_col)
        ml_letter   = get_column_letter(3)
        c = ws.cell(row=row_i, column=err_col,
                    value=f"=ABS({corr_letter}{row_i}-{ml_letter}{row_i})/{ml_letter}{row_i}*100")
        c.border = _BORDER; c.number_format = "0.00"

        ws.cell(row=row_i, column=sig_col, value="").border = _BORDER
        ws.cell(row=row_i, column=notes_col, value=s.notes).border = _BORDER

    _autofit(ws)


def _autofit(ws, min_width=8, max_width=30):
    for col in ws.columns:
        length = min_width
        for cell in col:
            if cell.value:
                length = max(length, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col[0].column)].width = length


# ═══════════════════════════════════════════════════════════════════════════════
# LOAD
# ═══════════════════════════════════════════════════════════════════════════════

def load_project(path: str) -> Project:
    wb = openpyxl.load_workbook(path, data_only=True)

    name, diameter_cm = "Imported Project", 1.27
    if "_xrf_meta" in wb.sheetnames:
        meta = wb["_xrf_meta"]
        for row in meta.iter_rows(values_only=True):
            if row[0] == "project_name" and row[1]:
                name = str(row[1])
            if row[0] == "diameter_cm" and row[1]:
                try: diameter_cm = float(row[1])
                except (ValueError, TypeError): pass

    project = Project(name=name, diameter_cm=diameter_cm)

    for sheet_name in wb.sheetnames:
        if sheet_name == "_xrf_meta":
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        title_cell = str(rows[0][0] or "").lower()
        is_validation = "validation" in title_cell

        # Read native config rows if present (rows 2–4: diameter, area, input mode)
        diam_override = _read_config_diameter(rows)
        input_mode    = _read_config_mode(rows) or INPUT_MASS

        result = None
        if "correction factor" in title_cell and not is_validation:
            result = _parse_native_calibration(rows, sheet_name)
        elif is_validation:
            result = _parse_native_analysis(rows, sheet_name)

        if result is None:
            result = _auto_detect_sheet(rows, sheet_name)

        # Collapse explicit-but-matches-default to "no override" for clean UI state
        if diam_override is not None and abs(diam_override - project.diameter_cm) < 1e-9:
            diam_override = None

        if isinstance(result, CalibrationSheet):
            result.diameter_cm = diam_override
            result.input_mode  = input_mode
            project.calibration_sheets.append(result)
        elif isinstance(result, AnalysisSheet):
            result.diameter_cm = diam_override
            result.input_mode  = input_mode
            project.analysis_sheets.append(result)

    return project


def _read_config_diameter(rows):
    """Look for 'Diameter (cm):' in column A and return the float in column B, or None."""
    for row in rows[:6]:
        if row and row[0] and "diameter" in str(row[0]).lower() and len(row) > 1:
            try: return float(row[1])
            except (ValueError, TypeError): return None
    return None


def _read_config_mode(rows):
    for row in rows[:6]:
        if row and row[0] and "input mode" in str(row[0]).lower() and len(row) > 1:
            v = str(row[1] or "").strip().lower()
            if v in (INPUT_MASS, INPUT_LOADING):
                return v
    return None


def _find_header_row(rows, keywords):
    for i, row in enumerate(rows):
        row_str = [str(c or "").lower() for c in row]
        if all(any(kw in cell for cell in row_str) for kw in keywords):
            return i
    return None


def _parse_native_calibration(rows, sheet_name):
    hdr_i = _find_header_row(rows, ["sample", "xrf"])
    if hdr_i is None:
        return None
    headers = [str(c or "").lower().strip() for c in rows[hdr_i]]
    try:
        id_col  = next(i for i, h in enumerate(headers) if "sample" in h)
        xrf_col = next(i for i, h in enumerate(headers) if "xrf" in h and "loading" in h)
        # The raw-input column: either "Mass (mg)" or "Mass Loading (mg/cm²)" — first column with "mass" or "loading" that isn't the computed loading
        mass_col = None
        for i, h in enumerate(headers):
            if (("mass" in h or "loading" in h)
                    and "xrf" not in h
                    and "σ" not in h and "unc" not in h
                    and i != id_col):
                mass_col = i; break
        if mass_col is None:
            return None
        unc_col = next((i for i, h in enumerate(headers) if "σ" in h or "unc" in h), None)
    except StopIteration:
        return None

    parts = sheet_name.split(" on ", 1)
    element   = parts[0].strip() if len(parts) == 2 else sheet_name.strip()
    substrate = parts[1].strip() if len(parts) == 2 else ""

    samples = []
    for row in rows[hdr_i + 1:]:
        if not row or row[id_col] is None and row[mass_col] is None:
            continue
        try:
            samples.append(CalibrationSample(
                sample_id=str(row[id_col] or ""),
                mass_mg=float(row[mass_col] or 0),
                xrf_loading=float(row[xrf_col] or 0),
                mass_uncertainty=float(row[unc_col] or 0.01) if unc_col is not None and row[unc_col] is not None else 0.01,
            ))
        except (ValueError, TypeError):
            continue
    return CalibrationSheet(element=element, substrate=substrate, samples=samples)


def _parse_native_analysis(rows, sheet_name):
    hdr_i = _find_header_row(rows, ["sample"])
    if hdr_i is None:
        return None
    headers = [str(c or "").lower().strip() for c in rows[hdr_i]]
    try:
        id_col   = next(i for i, h in enumerate(headers) if "sample" in h)
    except StopIteration:
        return None

    # Raw input column: first "mass" or "loading" column that isn't xrf/total/corr
    mass_col = None
    for i, h in enumerate(headers):
        if (("mass" in h or "loading" in h)
                and "xrf" not in h and "total" not in h and "corr" not in h
                and i != id_col):
            mass_col = i; break
    if mass_col is None:
        return None

    xrf_cols = {}
    for i, h in enumerate(headers):
        if h.startswith("xrf ") and "total" not in h:
            element = h.replace("xrf ", "").split(" ")[0]
            xrf_cols[element.capitalize()] = i
    if not xrf_cols:
        return None

    notes_col = next((i for i, h in enumerate(headers) if "note" in h), None)
    elements = list(xrf_cols.keys())

    samples = []
    for row in rows[hdr_i + 1:]:
        if not row or row[id_col] is None and row[mass_col] is None:
            continue
        try:
            loadings = {el: float(row[col] or 0) for el, col in xrf_cols.items()}
            notes = str(row[notes_col] or "") if notes_col is not None and row[notes_col] is not None else ""
            samples.append(AnalysisSample(
                sample_id=str(row[id_col] or ""),
                mass_mg=float(row[mass_col] or 0),
                xrf_loadings=loadings,
                notes=notes,
            ))
        except (ValueError, TypeError):
            continue
    return AnalysisSheet(name=sheet_name, elements=elements, samples=samples)


def _auto_detect_sheet(rows, sheet_name):
    """Loose detection for legacy files like Pb-on-Cu Corr. Factors.xlsx."""
    hdr_i = _find_header_row(rows, ["electrode", "mass"])
    if hdr_i is None:
        return None
    headers = [str(c or "").lower().strip() for c in rows[hdr_i]]

    xrf_element_cols = {}
    for i, h in enumerate(headers):
        if "xrf" in h and "loading" in h and "total" not in h and "corr" not in h:
            for w in h.replace("(", "").replace(")", "").split():
                if w not in ("xrf", "loading", "mg/cm²", "mg", "cm²") and len(w) <= 3:
                    xrf_element_cols[w.capitalize()] = i; break

    id_col   = next((i for i, h in enumerate(headers) if "electrode" in h or "sample" in h), 0)
    mass_col = next((i for i, h in enumerate(headers) if "mass" in h and "loading" not in h), 1)

    if len(xrf_element_cols) == 1:
        el = list(xrf_element_cols.keys())[0]
        xrf_col = xrf_element_cols[el]
        parts = sheet_name.split(" on ", 1)
        substrate = parts[1].strip() if len(parts) == 2 else ""
        samples = []
        for row in rows[hdr_i + 1:]:
            sid, mass, xrf = row[id_col], row[mass_col], row[xrf_col]
            if sid is None or mass is None: continue
            try:
                samples.append(CalibrationSample(
                    sample_id=str(sid), mass_mg=float(mass), xrf_loading=float(xrf or 0),
                ))
            except (ValueError, TypeError):
                continue
        if samples:
            return CalibrationSheet(element=el, substrate=substrate, samples=samples)

    elif len(xrf_element_cols) >= 2:
        elements = list(xrf_element_cols.keys())
        samples = []
        for row in rows[hdr_i + 1:]:
            sid, mass = row[id_col], row[mass_col]
            if sid is None or mass is None: continue
            try:
                loadings = {el: float(row[col] or 0) for el, col in xrf_element_cols.items()}
                samples.append(AnalysisSample(
                    sample_id=str(sid), mass_mg=float(mass), xrf_loadings=loadings,
                ))
            except (ValueError, TypeError):
                continue
        if samples:
            return AnalysisSheet(name=sheet_name, elements=elements, samples=samples)

    return None
